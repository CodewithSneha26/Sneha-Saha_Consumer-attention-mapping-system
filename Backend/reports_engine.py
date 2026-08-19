from database import SessionLocal
import models
from scoring_engine import calculate_shelf_scores
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

REPORTS_FOLDER = "generated_reports"
os.makedirs(REPORTS_FOLDER, exist_ok=True)


def gather_report_data():
    db = SessionLocal()

    attention_records = db.query(models.AttentionRecord).all()
    interactions = db.query(models.ProductInteraction).all()
    scores = calculate_shelf_scores()

    db.close()

    return {
        "attention_records": attention_records,
        "interactions": interactions,
        "scores": scores
    }


def generate_pdf_report():
    """Generates a consolidated PDF report - attention, engagement, shelf performance, conversion, marketing"""
    data = gather_report_data()
    filepath = os.path.join(REPORTS_FOLDER, "consumer_attention_report.pdf")

    doc = SimpleDocTemplate(filepath, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Consumer Attention Mapping System - Full Report", styles['Title']))
    elements.append(Spacer(1, 20))
    
    # Embed heatmap images
    from reportlab.platypus import Image as RLImage
    heatmap_files = [
        ("heatmap_1_store.png", "Store Presence Heatmap"),
        ("heatmap_2_shelves.png", "Shelf Dwell Time"),
        ("heatmap_3_product_attention.png", "Product Attention Matrix"),
        ("heatmap_4_traffic.png", "Customer Traffic Heatmap"),
    ]
    elements.append(Paragraph("Attention Heatmaps", styles['Heading2']))
    for filename, title in heatmap_files:
        if os.path.exists(filename):
            elements.append(Paragraph(title, styles['Heading3']))
            elements.append(RLImage(filename, width=400, height=280))
            elements.append(Spacer(1, 14))
    elements.append(Spacer(1, 10))

    # Section 1: Shelf Performance & Scoring
    elements.append(Paragraph("Shelf Performance & Attractiveness Scores", styles['Heading2']))
    table_data = [["Shelf", "Score", "Visibility", "Engagement", "Conversion", "Marketing"]]
    for shelf, d in data["scores"].items():
        table_data.append([
            shelf,
            str(d["attractiveness_score"]),
            str(d["shelf_visibility_score"]),
            str(d["engagement_score"]),
            str(d["conversion_potential_score"]),
            str(d["marketing_effectiveness_score"])
        ])

    table = Table(table_data, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#333333")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))

    # Section 2: Product Engagement Summary
    elements.append(Paragraph("Product Engagement Summary", styles['Heading2']))
    interaction_counts = {}
    for i in data["interactions"]:
        interaction_counts[i.interaction_type] = interaction_counts.get(i.interaction_type, 0) + 1

    engagement_data = [["Interaction Type", "Count"]]
    for itype, count in interaction_counts.items():
        engagement_data.append([itype, str(count)])

    engagement_table = Table(engagement_data, hAlign='LEFT')
    engagement_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#333333")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(engagement_table)
    elements.append(Spacer(1, 20))

    # Section 3: Consumer Attention Summary
    elements.append(Paragraph("Consumer Attention Summary", styles['Heading2']))
    total_attention_time = sum(r.duration_seconds for r in data["attention_records"])
    attentive_count = sum(1 for r in data["attention_records"] if r.attention_status == "Attentive")
    elements.append(Paragraph(f"Total recorded attention time: {total_attention_time} seconds", styles['Normal']))
    elements.append(Paragraph(f"Total attentive events: {attentive_count}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Section 4: Conversion Report
    elements.append(Paragraph("Conversion Report", styles['Heading2']))
    conversion_data = [["Shelf", "Total Interactions", "Purchased", "Conversion Rate (%)"]]
    for shelf, d in data["scores"].items():
        conversion_data.append([
            shelf,
            str(d["total_interactions"]),
            str(d["purchased_count"]),
            str(d["conversion_potential_score"])
        ])
    conversion_table = Table(conversion_data, hAlign='LEFT')
    conversion_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#333333")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(conversion_table)
    elements.append(Spacer(1, 20))

    # Section 5: Marketing Effectiveness Report
    elements.append(Paragraph("Marketing Effectiveness Report", styles['Heading2']))
    marketing_data = [["Shelf", "Compared Count", "Purchased Count", "Marketing Effectiveness (%)"]]
    for shelf, d in data["scores"].items():
        marketing_data.append([
            shelf,
            str(d["compared_count"]),
            str(d["purchased_count"]),
            str(d["marketing_effectiveness_score"])
        ])
    marketing_table = Table(marketing_data, hAlign='LEFT')
    marketing_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#333333")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(marketing_table)

    doc.build(elements)
    return filepath


def generate_excel_report():
    """Generates an Excel report with multiple sheets"""
    data = gather_report_data()
    filepath = os.path.join(REPORTS_FOLDER, "consumer_attention_report.xlsx")

    wb = openpyxl.Workbook()

    # Sheet 1: Shelf Scores
    ws1 = wb.active
    ws1.title = "Shelf Performance"
    headers = ["Shelf", "Attractiveness Score", "Visibility", "Engagement", "Conversion", "Marketing Effectiveness"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    for shelf, d in data["scores"].items():
        ws1.append([
            shelf,
            d["attractiveness_score"],
            d["shelf_visibility_score"],
            d["engagement_score"],
            d["conversion_potential_score"],
            d["marketing_effectiveness_score"]
        ])

    # Sheet 2: Product Interactions
    ws2 = wb.create_sheet("Product Interactions")
    ws2.append(["Person ID", "Shelf", "Interaction Type", "Duration (s)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    for i in data["interactions"]:
        ws2.append([i.person_track_id, i.shelf_zone, i.interaction_type, i.duration_seconds])

    # Sheet 3: Attention Records
    ws3 = wb.create_sheet("Attention Records")
    ws3.append(["Person ID", "Zone", "Attention Status", "Duration (s)"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    for r in data["attention_records"]:
        ws3.append([r.person_track_id, r.zone, r.attention_status, r.duration_seconds])

    # Sheet 4: Conversion Report
    ws4 = wb.create_sheet("Conversion Report")
    ws4.append(["Shelf", "Total Interactions", "Purchased", "Conversion Rate (%)"])
    for cell in ws4[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    for shelf, d in data["scores"].items():
        ws4.append([shelf, d["total_interactions"], d["purchased_count"], d["conversion_potential_score"]])

    # Section: Shopper Segment Breakdown
    from behavior_analysis import classify_shopper
    db_seg = SessionLocal()
    all_ids = db_seg.query(models.AttentionRecord.person_track_id).distinct().all()
    person_ids = [row[0] for row in all_ids]
    segment_counts = {}
    for pid in person_ids:
        result = classify_shopper(pid, db_seg)
        if isinstance(result, dict):
            segment_counts[result["segment"]] = segment_counts.get(result["segment"], 0) + 1
    db_seg.close()

    elements.append(Paragraph("Shopper Segment Breakdown", styles['Heading2']))
    segment_data = [["Shopper Segment", "Count"]]
    for seg, count in segment_counts.items():
        segment_data.append([seg, str(count)])
    if len(segment_data) > 1:
        segment_table = Table(segment_data, hAlign='LEFT')
        segment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#333333")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(segment_table)
    else:
        elements.append(Paragraph("No shopper segment data available yet.", styles['Normal']))
    elements.append(Spacer(1, 20))
    # Sheet 5: Marketing Effectiveness Report
    ws5 = wb.create_sheet("Marketing Effectiveness")
    ws5.append(["Shelf", "Compared Count", "Purchased Count", "Marketing Effectiveness (%)"])
    for cell in ws5[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    for shelf, d in data["scores"].items():
        ws5.append([shelf, d["compared_count"], d["purchased_count"], d["marketing_effectiveness_score"]])

    wb.save(filepath)
    return filepath


if __name__ == "__main__":
    pdf_path = generate_pdf_report()
    print(f"PDF report generated: {pdf_path}")

    excel_path = generate_excel_report()
    print(f"Excel report generated: {excel_path}")